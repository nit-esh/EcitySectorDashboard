#!/usr/bin/env python3
"""
update_regs.py — Reads ALL Numbers/Excel files in the Santhosha Data folder,
extracts registration counts, and updates the dashboard HTML with:
  1. LIVE_REGS    — upcoming batch-level counts (reg badges on programme cards)
  2. CENTRE_DATA  — annual IE/BSP/Shoonya/Samyama totals (Centre Insights page)
  3. MONTHLY_DATA — month-by-month breakdown for trend charts (where available)

Run this script from Terminal whenever any file in Santhosha Data is updated:
    python3 update_regs.py

Requires: macOS (uses osascript to export Numbers → CSV)
          pip install openpyxl  (for reading .xlsx directly)
"""

import subprocess, csv, os, re, json, sys, tempfile, glob, datetime

# openpyxl — for reading .xlsx directly (no AppleScript needed)
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE          = os.path.dirname(os.path.abspath(__file__))
HTML          = os.path.join(BASE, "isha_multi_centre_dashboard_v2.html")
SANTHOSHA_DIR = os.path.join(BASE, "Santhosha Data")
CURRENT_YEAR  = str(datetime.date.today().year)

# ── Normalisation patterns ────────────────────────────────────────────────────
CENTRE_PATTERNS = [
    (r'sadhguru sannidhi',                                    'sadhguru sannidhi'),
    (r'isha yoga cent(?:er|re)|iyc|coimbatore|velliangiri',   'isha yoga center'),
    (r'electronic city',                                      'electronic'),
    (r'kanakapura',                                           'kanakapura'),
    (r'sarjapur|sargapur',                                    'sarjapur'),
    (r'hsr layout|hsr',                                       'hsr'),
    (r'marathahalli|marathali',                               'marathahalli'),
    (r'malleswaram|malleshwaram',                             'malleswaram'),
    (r'vijayanagar|vijayanagara',                             'vijayanagar'),
    (r'jayanagar|jaynagar',                                   'jayanagar'),
    (r'banaswadi|banasawadi',                                 'banaswadi'),
    (r'hebbal|hebbala',                                       'hebbal'),
    (r'indiranagar|indira nagar',                             'indiranagar'),
    (r'mysuru|mysore',                                        'mysuru'),
    (r'hubballi|hubbali|hubli',                               'hubbali'),
    (r'ballari|bellary',                                      'ballari'),
    (r'belagavi|belgaum',                                     'belagavi'),
    (r'koramangala',                                          'koramangala'),
    (r'chikkaballapur',                                       'chikkaballapur'),
    # Sub-centres (primarily Monthly Satsang, parsed from Pivot Event xlsx)
    (r'girinagar',                                            'girinagar'),
    (r'yelahanka',                                            'yelahanka'),
    (r'chandapura',                                           'chandapura'),
    (r'begur',                                                'begur'),
    (r'budigere cross',                                       'budigere cross'),  # full name first
    (r'budigere',                                             'budigere cross'),  # short alias → same key
    (r'mangalore|mangaluru',                                  'mangalore'),
    (r'kengeri',                                              'kengeri'),
    (r'peenya',                                               'peenya'),
    (r'bg road|bannerghatta road',                            'bg road'),
    (r'tumkur',                                               'tumkur'),
    (r'whitefield',                                           'whitefield'),
    (r'singasandra',                                          'singasandra'),
]

PROG_PATTERNS = [
    (r'inner engineering|isha yoga|\bie\b',                          'inner engineering'),
    (r'bhava spandana|\bbsp\b',                                      'bhava spandana'),
    (r'shoonya',                                                     'shoonya intensive'),
    (r'samyama',                                                     'samyama'),
    (r'vairagya',                                                    'vairagya'),
    (r'eye care.*(upayoga|upa yoga)',                                 'eye care upa yoga'),
    (r'eye care.*(shanmukhi)',                                        'eye care shanmukhi'),
    (r'eye care',                                                    'eye care'),
    (r'shanmukhi',                                                   'shanmukhi'),
    (r'angamardana',                                                 'angamardana'),
    (r'surya kriya.*surya shakti|surya shakti.*surya kriya',          'surya kriya surya shakti'),
    (r'surya kriya',                                                 'surya kriya'),
    (r'surya shakti',                                                'surya shakti'),
    (r'yogasanas|yoga asana',                                        'yogasanas'),
    (r'bhuta shuddhi',                                               'bhuta shuddhi'),
    (r'bhastrika',                                                   'bhastrika'),
    (r'hatha yoga for children|hatha children',                      'hatha children'),
    (r'hatha yoga',                                                  'hatha yoga'),
    (r'thoppukarnam',                                                'thoppukarnam'),
    (r'jala neti',                                                   'jala neti'),
    (r'sutra neti',                                                  'sutra neti'),
    (r'nauli',                                                       'nauli'),
    (r'kapalbhati',                                                  'kapalbhati'),
    (r'trataka',                                                     'trataka'),
    (r'guru pooja|guru puja',                                        'guru pooja'),
    (r'upa yoga|upayoga',                                            'upa yoga'),
    (r'isha kriya',                                                  'isha kriya'),
    (r'chit shakti',                                                 'chit shakti'),
    (r'nada aradhana',                                               'nada aradhana'),
    (r'satsang|sathsang',                                            'satsang'),
    (r'isha janani',                                                 'isha janani'),
]

# Maps python centre key → CENTRE_DATA / MONTHLY_DATA key in HTML
CENTRE_KEY_MAP = {
    'electronic':    'Electronic City',
    'banaswadi':     'Banaswadi',
    'hebbal':        'Hebbal',
    'indiranagar':   'Indiranagar',
    'jayanagar':     'IP - Jayanagar',
    'malleswaram':   'IP - Malleswaram',
    'marathahalli':  'IP - Marathahalli',
    'vijayanagar':   'IP - Vijayanagar',
    'budigere cross':'Budigere Cross',
    'mangalore':     'Mangalore',
    'whitefield':    'Whitefield',
    'mysuru':        'Mysuru',
    'bg road':       'Bannerghatta Road',
    'ballari':       'Ballari',
    'belagavi':      'Belagavi',
    'hubbali':       'Hubbali',
    'koramangala':   'Koramangala',
    'kanakapura':    'Kanakapura Road',
    'girinagar':     'Girinagar',
    'hassan':        'Hassan',
    'tumkur':        'Tumkur',
    'udupi':         'Udupi',
}

# Maps normalised prog key → CENTRE_DATA / MONTHLY_DATA sub-key
PROG_TO_CD_KEY = {
    'inner engineering': 'ie',
    'bhava spandana':    'bsp',
    'shoonya intensive': 'shoonya',
    'samyama':           'samyama',
}

_MONTH_MAP = {
    'jan':'01','feb':'02','mar':'03','apr':'04','may':'05','jun':'06',
    'jul':'07','aug':'08','sep':'09','oct':'10','nov':'11','dec':'12'
}

def extract_start_date(text):
    m  = re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+(\d{1,2})\b', text.lower())
    ym = re.search(r'\b(20\d{2})\b', text)
    if m and ym:
        return f"{ym.group(1)}-{_MONTH_MAP[m.group(1)]}-{m.group(2).zfill(2)}"
    return None

def norm_centre(text):
    t = text.lower()
    for pat, key in CENTRE_PATTERNS:
        if re.search(pat, t): return key
    return None

def norm_prog(text):
    t = text.lower()
    for pat, key in PROG_PATTERNS:
        if re.search(pat, t): return key
    return None

# ── Find all eligible files recursively in Santhosha Data ────────────────────
def find_all_files():
    candidates = (
        glob.glob(os.path.join(SANTHOSHA_DIR, "**", "*.xlsx"),    recursive=True) +
        glob.glob(os.path.join(SANTHOSHA_DIR, "**", "*.numbers"), recursive=True) +
        glob.glob(os.path.join(SANTHOSHA_DIR, "**", "*.xls"),     recursive=True)
    )
    candidates = [f for f in candidates
                  if not os.path.basename(f).startswith('~$')
                  and not os.path.basename(f).startswith('.')]
    if not candidates:
        if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
            return [sys.argv[1]]
        print(f"✗ No Numbers/Excel files found in: {SANTHOSHA_DIR}")
        sys.exit(1)
    candidates.sort(key=os.path.getmtime, reverse=True)
    return candidates

# ── Export .numbers → CSV via AppleScript ────────────────────────────────────
def export_numbers_to_csv(numbers_path, tmp_csv):
    script = f'''
tell application "Numbers"
    open POSIX file "{numbers_path}"
    delay 2
    set theDoc to front document
    export theDoc to POSIX file "{tmp_csv}" as CSV
    close theDoc saving no
end tell
'''
    r = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    if r.returncode != 0:
        print(f"  ✗ AppleScript error: {r.stderr.strip()}")
        return False
    if not os.path.exists(tmp_csv):
        print(f"  ✗ CSV not created — is Numbers installed?")
        return False
    return True

# ── Detect and parse CRM "Total row" xlsx format ─────────────────────────────
# Format: Row1=labels, Row2=years, Row3="Count" headers, Row4="Total" with annual counts
# Row5+: detail rows — either "Month Year" (EC-style) or "Year" (other centres)
def try_parse_crm_xlsx(xlsx_path, regs, cd_updates, monthly_updates, source_label):
    """Returns True if file matches CRM format and was parsed; False to fall back."""
    if not HAS_OPENPYXL:
        return False
    try:
        wb   = openpyxl.load_workbook(xlsx_path, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
    except Exception as e:
        print(f"  ✗ openpyxl error: {e}")
        return False

    # Detect CRM format: row[1] has 4-digit year strings, row[3][0] == 'Total'
    if len(rows) < 4:
        return False
    years_row = rows[1]
    total_row = rows[3]
    if str(total_row[0]).strip().lower() != 'total':
        return False
    year_cols = [(i, str(v).strip()) for i, v in enumerate(years_row)
                 if v and re.match(r'^20\d{2}$|^1\d{3}$', str(v).strip())]
    if not year_cols:
        return False

    # Infer centre from parent folder name
    parent_folder = os.path.basename(os.path.dirname(xlsx_path))
    centre_key    = norm_centre(parent_folder)
    html_centre   = CENTRE_KEY_MAP.get(centre_key) if centre_key else None

    # Infer programme from filename (normalise underscores → spaces for pattern matching)
    fname    = os.path.basename(xlsx_path).lower().replace('_', ' ')
    prog_key = norm_prog(fname)
    cd_key   = PROG_TO_CD_KEY.get(prog_key) if prog_key else None

    if not html_centre or not cd_key:
        print(f"  ⚠ CRM format detected but couldn't infer centre ({parent_folder!r}) "
              f"or programme ({fname!r}) — skipping")
        return True  # still CRM format, just not updatable

    # Build col_index → year map for quick lookup
    col_to_year = {i: yr for i, yr in year_cols}

    # ── Annual totals → CENTRE_DATA ──────────────────────────────────────────
    # For IE: the Total row is programme-specific, use it directly.
    # For BSP/Shoonya/Samyama: the Total row is an all-member aggregate (same
    # value across all programme files), so we skip it and instead derive annual
    # totals by summing the programme-specific monthly rows (see section below).
    matched_annual = 0
    if cd_key == 'ie':
        for col_i, year in year_cols:
            val = total_row[col_i]
            if val is None:
                continue
            try:
                count = int(float(str(val)))
            except:
                continue
            if count <= 0:
                continue
            cd_updates.setdefault(html_centre, {}).setdefault(cd_key, {})[year] = count
            matched_annual += 1

    # ── Monthly breakdown → MONTHLY_DATA (only if row labels contain month names) ──
    # EC-style: "     January 2026", "     March 2026" etc.
    # Other:    "     2026" — year only, no monthly data available
    matched_monthly = 0
    month_name_re = re.compile(
        r'\b(january|february|march|april|may|june|july|august|september|october|november|december)\b',
        re.IGNORECASE
    )
    month_num = {
        'january':'01','february':'02','march':'03','april':'04',
        'may':'05','june':'06','july':'07','august':'08',
        'september':'09','october':'10','november':'11','december':'12'
    }

    for row in rows[4:]:
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        mn = month_name_re.search(label)
        if not mn:
            continue  # year-only row — no monthly breakdown available
        month_str = month_num[mn.group(1).lower()]
        # Extract year from label (e.g. "January 2026" → 2026)
        yr_m = re.search(r'\b(20\d{2})\b', label)
        if not yr_m:
            continue
        year = yr_m.group(1)

        # Sum ALL year columns for this row — the CRM pivot columns are member-join-year,
        # but the row label year is when the programme happened. Off-diagonal entries
        # (join year ≠ programme year) are the majority for repeat programmes like BSP/Shoonya.
        # Summing all columns gives the true participant count for that month.
        count = 0
        for col_i, col_yr in year_cols:
            val = row[col_i]
            if val is None:
                continue
            try:
                count += int(float(str(val)))
            except:
                pass
        if count <= 0:
            continue

        (monthly_updates
            .setdefault(html_centre, {})
            .setdefault(cd_key, {})
            .setdefault(year, {})[month_str]) = count
        matched_monthly += 1

    # ── Derive annual totals from monthly rows for BSP/Shoonya/Samyama ──────
    if cd_key != 'ie' and matched_monthly > 0:
        centre_monthly = monthly_updates.get(html_centre, {}).get(cd_key, {})
        for yr, months in centre_monthly.items():
            yr_total = sum(v for v in months.values() if v)
            if yr_total > 0:
                cd_updates.setdefault(html_centre, {}).setdefault(cd_key, {})[yr] = yr_total
                matched_annual += 1

    monthly_note = f", {matched_monthly} monthly entries" if matched_monthly else ""
    print(f"  ✓ {source_label} [CRM]: {matched_annual} annual{monthly_note} → "
          f"{html_centre} / {cd_key}")
    return True

# ── Read .xlsx via openpyxl → temp CSV (fallback for non-CRM xlsx) ───────────
def export_xlsx_to_csv(xlsx_path, tmp_csv):
    if not HAS_OPENPYXL:
        print(f"  ⚠ openpyxl not installed — falling back to AppleScript")
        return export_numbers_to_csv(xlsx_path, tmp_csv)
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        with open(tmp_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(['' if v is None else str(v) for v in row])
        return True
    except Exception as e:
        print(f"  ✗ openpyxl error: {e}")
        return False

# ── Route to correct exporter based on file type ─────────────────────────────
def export_to_csv(file_path, tmp_csv):
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        return export_xlsx_to_csv(file_path, tmp_csv)
    else:
        return export_numbers_to_csv(file_path, tmp_csv)

# ── Parse one CSV, merge into regs (LIVE_REGS) and cd_updates (CENTRE_DATA) ──
def parse_csv(csv_path, regs, cd_updates, source_label):
    skipped = 0
    matched = 0

    with open(csv_path, newline='', encoding='utf-8-sig', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            if not row: continue
            name = row[0].strip()
            count_str = ''
            for col in row[1:]:
                col = col.strip()
                if col and re.match(r'^-?\d+(\.\d+)?$', col):
                    count_str = col
                    break
            if not name or not count_str:
                continue
            try:
                count = int(float(count_str))
            except:
                continue
            if count < 0:
                continue

            centre = norm_centre(name)
            prog   = norm_prog(name)
            if not centre or not prog:
                skipped += 1
                continue

            # ── LIVE_REGS update ────────────────────────────────────────────
            date     = extract_start_date(name)
            prog_key = f"{prog}|{date}" if date else prog
            if centre not in regs:
                regs[centre] = {}
            if date:
                regs[centre][prog_key] = count
            else:
                regs[centre][prog_key] = regs[centre].get(prog_key, 0) + count

            # ── CENTRE_DATA update (IE / BSP / Shoonya / Samyama only) ─────
            cd_key      = PROG_TO_CD_KEY.get(prog)
            html_centre = CENTRE_KEY_MAP.get(centre)
            if cd_key and html_centre:
                year = (date[:4] if date else None) or CURRENT_YEAR
                if html_centre not in cd_updates:
                    cd_updates[html_centre] = {}
                if cd_key not in cd_updates[html_centre]:
                    cd_updates[html_centre][cd_key] = {}
                cd_updates[html_centre][cd_key][year] = (
                    cd_updates[html_centre][cd_key].get(year, 0) + count
                )

            matched += 1

    print(f"  ✓ {source_label}: {matched} matched, {skipped} skipped")
    return matched

# ── Read existing CENTRE_DATA from HTML ──────────────────────────────────────
CD_START = "/* __CENTRE_DATA_START__ */"
CD_END   = "/* __CENTRE_DATA_END__ */"

def _js_to_json(s):
    """Convert JS object literal (single-quoted keys, bare keys) to valid JSON."""
    # Quote bare JS keys (word chars before colon, not already quoted)
    s = re.sub(r"(?<!['\"\w])(\b[a-zA-Z_]\w*)\s*:", r'"\1":', s)
    # Replace single-quoted strings with double-quoted
    s = re.sub(r"'([^']*)'", r'"\1"', s)
    return s

def read_centre_data(html):
    m = re.search(re.escape(CD_START) + r'(.*?)' + re.escape(CD_END), html, re.DOTALL)
    if not m:
        return None, None
    block = m.group(1).strip()
    obj_m = re.search(r'const CENTRE_DATA\s*=\s*(\{.*\});', block, re.DOTALL)
    if not obj_m:
        return None, None
    try:
        return json.loads(obj_m.group(1)), block
    except json.JSONDecodeError:
        try:
            return json.loads(_js_to_json(obj_m.group(1))), block
        except json.JSONDecodeError as e:
            print(f"  ✗ Could not parse CENTRE_DATA: {e}")
            return None, None

# ── Read existing MONTHLY_DATA from HTML ─────────────────────────────────────
MD_START = "/* __MONTHLY_DATA_START__ */"
MD_END   = "/* __MONTHLY_DATA_END__ */"

def read_monthly_data(html):
    m = re.search(re.escape(MD_START) + r'(.*?)' + re.escape(MD_END), html, re.DOTALL)
    if not m:
        return None
    block = m.group(1)
    js = re.search(r'const MONTHLY_DATA\s*=\s*(\{.*\});', block, re.DOTALL)
    if not js:
        return None
    try:
        return json.loads(js.group(1))
    except json.JSONDecodeError:
        try:
            return json.loads(_js_to_json(js.group(1)))
        except json.JSONDecodeError as e:
            print(f"  ✗ Could not parse MONTHLY_DATA: {e}")
            return None

# ── Build JS object string for MONTHLY_DATA (compact inner dicts) ────────────
def monthly_data_to_js(md):
    """Serialize MONTHLY_DATA back to compact JS object literal."""
    lines = ['const MONTHLY_DATA = {']
    centres = sorted(md.keys())
    for ci, centre in enumerate(centres):
        comma_c = ',' if ci < len(centres) - 1 else ''
        lines.append(f"  '{centre}': {{")
        progs = md[centre]
        prog_keys = list(progs.keys())
        for pi, prog in enumerate(prog_keys):
            comma_p = ',' if pi < len(prog_keys) - 1 else ''
            lines.append(f"    {prog}: {{")
            years = sorted(progs[prog].keys())
            for yi, year in enumerate(years):
                months = progs[prog][year]
                comma_y = ',' if yi < len(years) - 1 else ''
                inner = ','.join(f"'{m}':{v}" for m, v in sorted(months.items()))
                lines.append(f"      '{year}':{{{inner}}}{comma_y}")
            lines.append(f"    }}{comma_p}")
        lines.append(f"  }}{comma_c}")
    lines.append('};')
    return '\n'.join(lines)

# ── Inject LIVE_REGS, CENTRE_DATA, and MONTHLY_DATA into HTML ────────────────
LIVE_START = "/* __LIVE_REGS_START__ */"
LIVE_END   = "/* __LIVE_REGS_END__ */"

def inject_html(html, regs, centre_data, monthly_data, latest_mtime):
    # 1. LIVE_REGS block
    dt          = datetime.datetime.fromtimestamp(latest_mtime)
    updated_str = dt.strftime('%d %b %Y, %I:%M %p')
    regs_json   = json.dumps(regs, indent=2, ensure_ascii=False)
    live_block  = (f"{LIVE_START}\nconst LIVE_REGS = {regs_json};\n"
                   f"const LIVE_REGS_UPDATED = {json.dumps(updated_str)};\n{LIVE_END}")

    if LIVE_START not in html:
        print("  ✗ LIVE_REGS markers not found in HTML")
        sys.exit(1)
    html = re.sub(re.escape(LIVE_START) + r'.*?' + re.escape(LIVE_END),
                  live_block, html, flags=re.DOTALL)
    print(f"  ✓ Updated LIVE_REGS  (as of {updated_str})")

    # 2. CENTRE_DATA block
    if centre_data is not None and CD_START in html:
        cd_json  = json.dumps(centre_data, indent=2, ensure_ascii=False)
        cd_block = f"{CD_START}\nconst CENTRE_DATA = {cd_json};\n{CD_END}"
        html = re.sub(re.escape(CD_START) + r'.*?' + re.escape(CD_END),
                      cd_block, html, flags=re.DOTALL)
        print(f"  ✓ Updated CENTRE_DATA")
    else:
        print(f"  ⚠ CENTRE_DATA markers not found — skipping")

    # 3. MONTHLY_DATA block
    if monthly_data is not None and MD_START in html:
        md_js    = monthly_data_to_js(monthly_data)
        md_block = f"{MD_START}\n// Monthly data — centres with month-level CRM data (others have annual totals only)\n{md_js}\n{MD_END}"
        html = re.sub(re.escape(MD_START) + r'.*?' + re.escape(MD_END),
                      md_block, html, flags=re.DOTALL)
        print(f"  ✓ Updated MONTHLY_DATA")
    else:
        print(f"  ⚠ MONTHLY_DATA markers not found — skipping")

    with open(HTML, 'w', encoding='utf-8') as f:
        f.write(html)

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("\n── Karnataka Registration Counts Updater ──")

    files = find_all_files()
    print(f"  Found {len(files)} file(s) in Santhosha Data:")
    for f in files:
        print(f"    • {os.path.relpath(f, SANTHOSHA_DIR)}")

    # Read current HTML once
    with open(HTML, 'r', encoding='utf-8') as f:
        html = f.read()

    # Read existing CENTRE_DATA and MONTHLY_DATA
    centre_data, _ = read_centre_data(html)
    if centre_data is None:
        print("  ⚠ Could not parse CENTRE_DATA from HTML — will skip Centre Insights update")

    monthly_data = read_monthly_data(html)
    if monthly_data is None:
        print("  ⚠ Could not parse MONTHLY_DATA from HTML — will skip monthly chart update")

    regs            = {}   # merged LIVE_REGS
    cd_updates      = {}   # CENTRE_DATA changes to apply
    monthly_updates = {}   # MONTHLY_DATA changes to apply
    tmp_files       = []
    latest_mtime    = max(os.path.getmtime(f) for f in files)

    for xlsx in files:
        label = os.path.relpath(xlsx, SANTHOSHA_DIR)
        ext   = os.path.splitext(xlsx)[1].lower()

        # Try CRM "Total row" format first (xlsx only — no CSV export needed)
        if ext in ('.xlsx', '.xls') and HAS_OPENPYXL:
            if try_parse_crm_xlsx(xlsx, regs, cd_updates, monthly_updates, label):
                continue   # handled — skip CSV export path

        # Standard batch-level format (Numbers files + non-CRM xlsx)
        if ext == '.numbers' and sys.platform != 'darwin':
            print(f"  ⚠ Skipping {label} (.numbers requires macOS)")
            continue
        tmp_csv = tempfile.mktemp(suffix=".csv")
        tmp_files.append(tmp_csv)
        print(f"\nExporting {label} → CSV …")
        if export_to_csv(xlsx, tmp_csv):
            parse_csv(tmp_csv, regs, cd_updates, label)
        else:
            print(f"  ⚠ Skipping {label} (export failed)")

    # Apply cd_updates → patch CENTRE_DATA
    if centre_data and cd_updates:
        changed = []
        for html_centre, progs in cd_updates.items():
            centre_data.setdefault(html_centre, {})
            for cd_key, year_counts in progs.items():
                centre_data[html_centre].setdefault(cd_key, {})
                for year, count in year_counts.items():
                    old = centre_data[html_centre][cd_key].get(year, 0)
                    centre_data[html_centre][cd_key][year] = count
                    if old != count:
                        changed.append(f"  {html_centre}/{cd_key}/{year}: {old} → {count}")
        if changed:
            print(f"\nCENTRE_DATA changes:")
            for c in changed: print(c)
        else:
            print(f"\nCENTRE_DATA: no changes")

    # Apply monthly_updates → patch MONTHLY_DATA
    if monthly_data and monthly_updates:
        changed = []
        for html_centre, progs in monthly_updates.items():
            monthly_data.setdefault(html_centre, {})
            for cd_key, year_months in progs.items():
                monthly_data[html_centre].setdefault(cd_key, {})
                for year, months in year_months.items():
                    monthly_data[html_centre][cd_key].setdefault(year, {})
                    for month, count in months.items():
                        old = monthly_data[html_centre][cd_key][year].get(month, 0)
                        monthly_data[html_centre][cd_key][year][month] = count
                        if old != count:
                            changed.append(f"  {html_centre}/{cd_key}/{year}-{month}: {old} → {count}")
        if changed:
            print(f"\nMONTHLY_DATA changes:")
            for c in changed: print(c)
        else:
            print(f"\nMONTHLY_DATA: no changes")

    # Print LIVE_REGS summary
    total = sum(sum(p.values()) for p in regs.values())
    print(f"\n{'─'*62}")
    print(f"  {'CENTRE':<22}  {'PROGRAMME':<28}  {'COUNT':>5}")
    print(f"{'─'*62}")
    for centre in sorted(regs):
        for prog_key, count in sorted(regs[centre].items()):
            if '|' in prog_key:
                prog, date = prog_key.split('|', 1)
                prog_display = f"{prog.title()} ({date})"
            else:
                prog_display = prog_key.title()
            print(f"  {centre:<22}  {prog_display:<28}  {count:>5}")
    print(f"{'─'*62}")
    print(f"  {'TOTAL':<52}  {total:>5}")

    print(f"\nInjecting into dashboard …")
    inject_html(
        html,
        regs,
        centre_data if cd_updates else None,
        monthly_data if monthly_updates else None,
        latest_mtime
    )

    # Cleanup
    for t in tmp_files:
        try: os.remove(t)
        except: pass

    print(f"\n✓ Done — {len(files)} file(s) processed.\n")
